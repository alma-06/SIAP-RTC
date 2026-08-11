from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook


REQUIRED_COLUMNS = (
    "Pauta de transmisión",
    "Estado",
    "Tiempo Fiscal",
    "Canal Base",
    "Orden",
    "Fecha",
    "Dependencia CAM. SEN.",
    "Clave",
    "Campaña",
    "Versión",
)


@dataclass(frozen=True)
class IngestedRecord:
    source_file: str
    values: dict[str, object]


@dataclass(frozen=True)
class IngestResult:
    records: tuple[IngestedRecord, ...]
    source_files: tuple[str, ...]
    warnings: tuple[str, ...]


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text.strip())
    return text.casefold()


def _find_headers(row: Iterable[object]) -> dict[str, int]:
    normalized = {_normalize_header(value): index for index, value in enumerate(row)}
    return {
        required: normalized[_normalize_header(required)]
        for required in REQUIRED_COLUMNS
        if _normalize_header(required) in normalized
    }


def _is_senado(value: object) -> bool:
    normalized = _normalize_header(value)
    return normalized in {
        "cam. sen.",
        "cam sen",
        "camara de senadores",
        "cámara de senadores",
    }


def ingest_rtc_workbooks(paths: Iterable[str | Path]) -> IngestResult:
    records: list[IngestedRecord] = []
    warnings: list[str] = []
    source_files: list[str] = []

    for raw_path in paths:
        path = Path(raw_path)
        source_files.append(path.name)
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                warnings.append(f"Hoja vacía: {path.name}/{sheet.title}")
                continue
            headers = _find_headers(header_row)
            missing = [column for column in REQUIRED_COLUMNS if column not in headers]
            if missing:
                warnings.append(
                    f"Columnas faltantes en {path.name}/{sheet.title}: {', '.join(missing)}"
                )
                continue
            for row in rows:
                if not any(value is not None for value in row):
                    continue
                if not _is_senado(row[headers["Dependencia CAM. SEN."]]):
                    continue
                values = {
                    column: row[headers[column]]
                    for column in REQUIRED_COLUMNS
                }
                records.append(IngestedRecord(path.name, values))
        workbook.close()

    return IngestResult(tuple(records), tuple(source_files), tuple(warnings))
