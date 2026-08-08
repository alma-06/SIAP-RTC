from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


CANONICAL_HEADERS = {
    "pauta de transmisión": "pauta_transmision",
    "pauta de transmision": "pauta_transmision",
    "estado": "estado",
    "tiempo fiscal": "tiempo_fiscal",
    "canal base": "canal_base",
    "orden": "orden",
    "fecha": "fecha",
    "dependencia cam. sen.": "dependencia_cam_sen",
    "dependencia cam sen": "dependencia_cam_sen",
    "clave": "clave",
    "campaña": "campana",
    "campana": "campana",
    "versión": "version",
    "version": "version",
}


def _header_key(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.strip())
    return value


@dataclass(frozen=True)
class NormalizedRTCRecord:
    values: dict[str, object]
    source_row: int
    source_sheet: str


class RTCExcelReader:
    def read(self, path: Path) -> list[NormalizedRTCRecord]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            records: list[NormalizedRTCRecord] = []
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                header_values = next(rows, ())
                mapping = self._build_mapping(header_values)
                if not mapping:
                    continue
                for row_number, row in enumerate(rows, start=2):
                    if not any(value is not None and str(value).strip() for value in row):
                        continue
                    values = {
                        canonical: normalize_value(row[index] if index < len(row) else None)
                        for index, canonical in mapping.items()
                    }
                    records.append(
                        NormalizedRTCRecord(
                            values=values,
                            source_row=row_number,
                            source_sheet=sheet.title,
                        )
                    )
            return records
        finally:
            workbook.close()

    @staticmethod
    def _build_mapping(headers: tuple[object, ...]) -> dict[int, str]:
        mapping: dict[int, str] = {}
        for index, header in enumerate(headers):
            canonical = CANONICAL_HEADERS.get(_header_key(header))
            if canonical:
                mapping[index] = canonical
        return mapping
