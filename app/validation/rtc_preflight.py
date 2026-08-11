from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook

from app.pipeline.rtc_ingest import REQUIRED_COLUMNS


@dataclass(frozen=True)
class SheetValidation:
    sheet: str
    header_row: int | None
    missing_columns: tuple[str, ...]
    data_rows: int
    senate_rows: int
    blank_rows: int
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookValidation:
    filename: str
    sheets: tuple[SheetValidation, ...]
    warnings: tuple[str, ...]

    @property
    def valid(self) -> bool:
        return bool(self.sheets) and not any(
            sheet.missing_columns for sheet in self.sheets if sheet.header_row is not None
        )


def _norm(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.strip()).casefold()


def _senado(value: object) -> bool:
    return _norm(value) in {
        "cam. sen.",
        "cam sen",
        "camara de senadores",
        "cámara de senadores",
    }


def validate_rtc_workbook(path: str | Path) -> WorkbookValidation:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet_results: list[SheetValidation] = []
    workbook_warnings: list[str] = []
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                sheet_results.append(SheetValidation(sheet.title, None, tuple(REQUIRED_COLUMNS), 0, 0, 0, ("Hoja vacía",)))
                continue

            positions = {_norm(value): idx for idx, value in enumerate(header)}
            missing = tuple(column for column in REQUIRED_COLUMNS if _norm(column) not in positions)
            if missing:
                sheet_results.append(SheetValidation(sheet.title, 1, missing, 0, 0, 0, ("Estructura incompatible con el esquema RTC esperado",)))
                continue

            data_rows = senate_rows = blank_rows = 0
            warnings: list[str] = []
            dependency_index = positions[_norm("Dependencia CAM. SEN.")]
            for row in rows:
                if not any(value is not None for value in row):
                    blank_rows += 1
                    continue
                data_rows += 1
                if _senado(row[dependency_index] if dependency_index < len(row) else None):
                    senate_rows += 1

            if data_rows and senate_rows == 0:
                warnings.append("No se encontraron registros identificables de Cámara de Senadores")
            sheet_results.append(SheetValidation(sheet.title, 1, missing, data_rows, senate_rows, blank_rows, tuple(warnings)))
    finally:
        workbook.close()

    if not sheet_results:
        workbook_warnings.append("El libro no contiene hojas")
    return WorkbookValidation(source.name, tuple(sheet_results), tuple(workbook_warnings))


def validate_rtc_workbooks(paths: Iterable[str | Path]) -> tuple[WorkbookValidation, ...]:
    return tuple(validate_rtc_workbook(path) for path in paths)
