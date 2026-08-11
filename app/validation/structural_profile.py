from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.export.package_manifest import _hash_file
from app.validation.rtc_preflight import REQUIRED_COLUMNS


@dataclass(frozen=True)
class SheetProfile:
    name: str
    max_row: int
    max_column: int
    headers: tuple[str, ...]
    missing_columns: tuple[str, ...]
    data_rows: int
    senate_rows: int
    blank_rows: int
    date_values: tuple[str, ...] = ()
    time_fiscal_samples: tuple[str, ...] = ()


@dataclass(frozen=True)
class WorkbookStructuralProfile:
    path: str
    sha256: str
    sheet_count: int
    sheets: tuple[SheetProfile, ...] = field(default_factory=tuple)


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _is_senate(value: object) -> bool:
    return _text(value).casefold() in {"cam. sen.", "cam sen", "cámara de senadores", "camara de senadores"}


def profile_workbook(path: str | Path) -> WorkbookStructuralProfile:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    profiles: list[SheetProfile] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, ())
        headers = tuple(_text(value) for value in header_row)
        positions = {header.casefold(): index for index, header in enumerate(headers)}
        missing = tuple(column for column in REQUIRED_COLUMNS if column.casefold() not in positions)
        data_rows = senate_rows = blank_rows = 0
        dates: list[str] = []
        times: list[str] = []
        for row in rows:
            values = tuple(row)
            if not any(value is not None and _text(value) for value in values):
                blank_rows += 1
                continue
            data_rows += 1
            dep_index = positions.get("dependencia cam. sen.")
            if dep_index is not None and dep_index < len(values) and _is_senate(values[dep_index]):
                senate_rows += 1
            date_index = positions.get("fecha")
            if date_index is not None and date_index < len(values) and values[date_index] is not None and len(dates) < 5:
                dates.append(_text(values[date_index]))
            time_index = positions.get("tiempo fiscal")
            if time_index is not None and time_index < len(values) and values[time_index] is not None and len(times) < 5:
                times.append(_text(values[time_index]))
        profiles.append(SheetProfile(sheet.title, sheet.max_row, sheet.max_column, headers, missing, data_rows, senate_rows, blank_rows, tuple(dates), tuple(times)))
    workbook.close()
    return WorkbookStructuralProfile(str(source), _hash_file(source), len(profiles), tuple(profiles))
