from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import hashlib

from openpyxl import load_workbook


REQUIRED_HEADERS = {
    "Pauta de transmisión",
    "Estado",
    "Tiempo Fiscal",
    "Canal Base",
    "Orden",
    "Fecha",
    "Dependencia CAM. SEN.",
    "Clave",
    "Campaña",
    "Versión",
}


@dataclass(frozen=True)
class FileValidationResult:
    path: Path
    size_bytes: int
    sha256: str
    valid: bool
    headers: tuple[str, ...] = ()
    missing_headers: tuple[str, ...] = ()
    error: str | None = None


@dataclass
class PreflightValidator:
    required_headers: set[str] = field(default_factory=lambda: set(REQUIRED_HEADERS))

    def validate(self, path: Path) -> FileValidationResult:
        if not path.exists():
            return FileValidationResult(path, 0, "", False, error="El archivo ya no existe.")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return FileValidationResult(path, path.stat().st_size, "", False, error="Formato no compatible; se requiere XLSX o XLSM.")
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = tuple(str(value).strip() for value in values if value is not None)
            missing = tuple(sorted(self.required_headers - set(headers)))
            workbook.close()
        except Exception as exc:
            return FileValidationResult(path, path.stat().st_size, digest, False, error=f"No fue posible leer el libro: {exc}")
        return FileValidationResult(
            path=path,
            size_bytes=path.stat().st_size,
            sha256=digest,
            valid=not missing,
            headers=headers,
            missing_headers=missing,
            error=None if not missing else "Faltan encabezados requeridos.",
        )
