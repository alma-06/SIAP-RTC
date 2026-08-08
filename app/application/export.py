"""Executive Excel export for SIAP-RTC historical records."""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from app.infrastructure.orm import RtcRecordModel

HEADERS = [
    "Pauta de transmisión", "Estado", "Tiempo Fiscal", "Canal Base", "Orden",
    "Fecha", "Dependencia (CAM. SEN.)", "Clave", "Campaña", "Versión",
]


class RtcExcelExporter:
    """Generate a self-contained executive workbook from RTC records."""

    def export(self, records: Sequence[RtcRecordModel], output: Path) -> Path:
        workbook = Workbook()
        data = workbook.active
        data.title = "Base consolidada"
        data.append(HEADERS)
        for record in records:
            data.append([
                record.pauta_transmision, record.estado, record.tiempo_fiscal,
                record.canal_base, record.orden, record.fecha, record.dependencia,
                record.clave, record.campana, record.version,
            ])
        end_row = max(2, data.max_row)
        table = Table(displayName="TablaRTC", ref=f"A1:J{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        data.add_table(table)
        data.freeze_panes = "A2"
        for column in range(1, 11):
            width = max(14, min(45, max(len(str(data.cell(r, column).value or "")) for r in range(1, end_row + 1)) + 2))
            data.column_dimensions[get_column_letter(column)].width = width

        summary = workbook.create_sheet("Resumen ejecutivo")
        summary.append(["Indicador", "Valor"])
        summary.append(["Total de spots", len(records)])
        summary.append(["Campañas", len({r.campana for r in records})])
        summary.append(["Versiones", len({r.version for r in records})])
        summary.append(["Canales", len({r.canal_base for r in records})])
        summary.append(["Estados", len({r.estado for r in records})])
        summary.freeze_panes = "A2"

        campaign = workbook.create_sheet("Por campaña")
        campaign.append(["Campaña", "Spots"])
        counts: dict[str, int] = {}
        for record in records:
            counts[record.campana] = counts.get(record.campana, 0) + 1
        for name, count in sorted(counts.items()):
            campaign.append([name, count])
        if campaign.max_row >= 2:
            chart = BarChart()
            chart.title = "Spots por campaña"
            chart.y_axis.title = "Spots"
            chart.x_axis.title = "Campaña"
            chart.add_data(Reference(campaign, min_col=2, min_row=1, max_row=campaign.max_row), titles_from_data=True)
            chart.set_categories(Reference(campaign, min_col=1, min_row=2, max_row=campaign.max_row))
            summary.add_chart(chart, "D2")

        workbook.save(output)
        return output
