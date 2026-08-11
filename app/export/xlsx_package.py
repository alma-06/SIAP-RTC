from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.pipeline.integrated import IntegratedPipelineResult
from app.export.executive_package import build_executive_summary


def write_xlsx_package(result: IntegratedPipelineResult, path: str | Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    data = build_executive_summary(result)
    for row, (key, value) in enumerate(data.items(), start=1):
        summary.cell(row=row, column=1, value=key)
        summary.cell(row=row, column=2, value=str(value))

    base = workbook.create_sheet("Base Consolidada")
    headers = sorted({key for item in result.consolidation.records for key in item.record.values})
    if headers:
        base.append(["Periodo", "Archivo fuente", *headers])
        for item in result.consolidation.records:
            base.append([item.period, item.source_file, *[item.record.values.get(h, "") for h in headers]])

    if result.reconciliation is not None:
        sheet = workbook.create_sheet("Conciliación")
        sheet.append(["Clave", "Clasificación", "Fuente anterior", "Fuente actual", "Campos modificados"])
        for decision in result.reconciliation.decisions:
            sheet.append([
                decision.key, decision.classification, decision.previous_source or "",
                decision.current_source or "", ", ".join(decision.changed_fields),
            ])

    if result.indicators is not None:
        sheet = workbook.create_sheet("Indicadores")
        for row, (key, value) in enumerate(asdict(result.indicators).items(), start=1):
            sheet.cell(row=row, column=1, value=key)
            sheet.cell(row=row, column=2, value=value)

    sources = workbook.create_sheet("Fuentes")
    sources.append(["Archivo", "Periodo", "SHA-256"])
    for source in result.evidence.sources:
        sources.append([source.filename, source.period, source.sha256])

    warnings = workbook.create_sheet("Advertencias")
    warnings.append(["Advertencia"])
    for warning in result.evidence.warnings:
        warnings.append([warning])

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions if sheet.max_row > 1 else None

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
