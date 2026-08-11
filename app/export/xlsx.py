from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from app.export.contracts import ExportPayload


def export_xlsx(payload: ExportPayload, output_path: str | Path) -> Path:
    path = Path(output_path)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen ejecutivo"

    summary["A1"] = payload.title
    summary["A1"].font = Font(bold=True, size=16)
    summary["A2"] = payload.subtitle
    summary["A4"] = "Periodo"
    summary["B4"] = payload.period
    summary["A5"] = "Estado"
    summary["B5"] = payload.status
    summary["A7"] = "Indicador"
    summary["B7"] = "Valor"
    summary["A7"].font = Font(bold=True)
    summary["B7"].font = Font(bold=True)

    row = 8
    for label, value in payload.metrics.items():
        summary.cell(row=row, column=1, value=label)
        summary.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    summary.cell(row=row, column=1, value="Alertas")
    summary.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    if payload.alerts:
        for alert in payload.alerts:
            summary.cell(row=row, column=1, value=alert)
            row += 1
    else:
        summary.cell(row=row, column=1, value="Sin alertas")
        row += 1

    row += 1
    summary.cell(row=row, column=1, value="Evidencia")
    summary.cell(row=row, column=2, value=payload.evidence_id)

    indicators = workbook.create_sheet("Indicadores")
    indicators.append(["Indicador", "Valor"])
    for cell in indicators[1]:
        cell.font = Font(bold=True)
    for label, value in payload.metrics.items():
        indicators.append([label, value])

    alerts = workbook.create_sheet("Alertas")
    alerts.append(["Periodo", "Alerta"])
    for cell in alerts[1]:
        cell.font = Font(bold=True)
    for message in payload.alerts or ("Sin alertas",):
        alerts.append([payload.period, message])

    evidence = workbook.create_sheet("Evidencia")
    evidence.append(["Campo", "Valor"])
    for cell in evidence[1]:
        cell.font = Font(bold=True)
    evidence.append(["Periodo", payload.period])
    evidence.append(["Estado", payload.status])
    evidence.append(["Evidence ID", payload.evidence_id])

    comparison = workbook.create_sheet("Comparativo")
    comparison.append(["Periodo", "Estado", "Evidencia"])
    for cell in comparison[1]:
        cell.font = Font(bold=True)
    comparison.append([payload.period, payload.status, payload.evidence_id])

    for sheet in workbook.worksheets:
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 28
        if sheet.max_column >= 3:
            sheet.column_dimensions["C"].width = 28

    workbook.save(path)
    return path
