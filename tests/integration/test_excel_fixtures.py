from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.fixtures.generate_fixtures import generate


@pytest.fixture(scope="session")
def fixture_dir(tmp_path_factory: pytest.TempPathFactory) -> Path:
    root = tmp_path_factory.mktemp("rtc-fixtures")
    generate(root)
    return root


def test_valid_fixture_is_readable(fixture_dir: Path) -> None:
    workbook = load_workbook(fixture_dir / "valid_single_sheet.xlsx", read_only=True)
    sheet = workbook["Pauta"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "Pauta de transmisión"
    assert len(rows) == 3


def test_duplicate_fixture_contains_duplicate_input(fixture_dir: Path) -> None:
    workbook = load_workbook(fixture_dir / "duplicate_rows.xlsx", read_only=True)
    rows = list(workbook["Pauta"].iter_rows(values_only=True))
    assert rows[1] == rows[3]
