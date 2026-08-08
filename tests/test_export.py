from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from app.application.export import RtcExcelExporter
from app.infrastructure.orm import RtcRecordModel


def test_export_creates_consolidated_and_summary_sheets(tmp_path: Path) -> None:
    records = [
        RtcRecordModel(
            id="1", batch_id="b", pauta_transmision="P", estado="VIGENTE",
            tiempo_fiscal="FISCAL", canal_base="CANAL", orden="O", fecha=date(2026, 8, 1),
            dependencia="CAM. SEN.", clave="K", campana="Campaña A", version="V1",
        )
    ]
    output = tmp_path / "reporte.xlsx"
    RtcExcelExporter().export(records, output)
    workbook = load_workbook(output)
    assert "Base consolidada" in workbook.sheetnames
    assert "Resumen ejecutivo" in workbook.sheetnames
    assert workbook["Base consolidada"].max_row == 2
    assert workbook["Resumen ejecutivo"]["B2"].value == 1
