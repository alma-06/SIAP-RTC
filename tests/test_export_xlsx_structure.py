from openpyxl import load_workbook

from app.dashboard.fact_sheet import build_fact_sheet
from app.dashboard.model import DashboardPeriod
from app.export.contracts import build_export_payload
from app.export.xlsx import export_xlsx


def test_xlsx_contains_executive_and_traceability_sheets(tmp_path) -> None:
    period = DashboardPeriod("2026-Q2", 100, 10, 5, 80, 5, 0.80, 0.05, 0.10, 0.05, True, (), "EV-01")
    payload = build_export_payload(build_fact_sheet(period))
    output = export_xlsx(payload, tmp_path / "siap_rtc.xlsx")
    workbook = load_workbook(output, read_only=True, data_only=True)
    assert workbook.sheetnames == ["Resumen ejecutivo", "Indicadores", "Alertas", "Evidencia", "Comparativo"]
    assert workbook["Evidencia"]["B3"].value == "EV-01"
    assert workbook["Comparativo"]["A2"].value == "2026-Q2"
