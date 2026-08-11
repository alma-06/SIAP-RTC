from openpyxl import load_workbook, Workbook

from app.export.xlsx_package import write_xlsx_package
from app.pipeline.integrated import run_integrated_period

HEADERS = ["Pauta de transmisión", "Estado", "Tiempo Fiscal", "Canal Base", "Orden", "Fecha", "Dependencia CAM. SEN.", "Clave", "Campaña", "Versión"]


def test_write_xlsx_package_creates_expected_sheets(tmp_path) -> None:
    source = tmp_path / "rtc.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["Pauta", "Programada", "00:00:30", "AM", "O1", "11/08/2026", "CAM. SEN.", "C1", "Campaña", "V1"])
    workbook.save(source)

    result = run_integrated_period([source], "2026-Q2")
    output = tmp_path / "SIAP-RTC.xlsx"
    write_xlsx_package(result, output)
    generated = load_workbook(output, read_only=True)
    assert generated.sheetnames == ["Resumen", "Base Consolidada", "Fuentes", "Advertencias"]
